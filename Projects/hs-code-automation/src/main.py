from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
import re
import json

# ======================================
# CONFIGURATION
# ======================================

INSW_URL = "https://insw.go.id/intr"

FILE_PATH = "./data/sample_hs_codes.xlsx"

OUTPUT_JSON = "./output/hasil_hs.json"

HEADLESS = True

# ======================================
# SCRAPING FUNCTION
# ======================================

def get_hs_info(page, hs_code):

    result = {
        "hs_code": hs_code,
        "bm_mfn": None,
        "ppn": None,
        "pph_api": None,
        "pph_non_api": None,
        "izin": []
    }

    page.goto(INSW_URL)

    page.wait_for_timeout(3000)

    # Search HS Code

    page.get_by_placeholder(
        "Cari kode HS / Uraian HS"
    ).fill(hs_code)

    page.keyboard.press("Enter")

    page.wait_for_timeout(3000)

    # Open detail page

    page.get_by_role(
        "button",
        name="Detail"
    ).click()

    page.wait_for_timeout(3000)

    # Open import regulation section

    page.get_by_text(
        "Regulasi Impor (Lartas Border)"
    ).click()

    page.wait_for_timeout(1000)

    # Open BC 2.0 card

    page.locator(
        "span.card-title",
        has_text="BC 2.0"
    ).first.click()

    page.wait_for_timeout(2000)

    text = page.locator("body").inner_text()

    # ======================================
    # BM MFN
    # ======================================

    m = re.search(
        r"BM MFN.*?:\s*([\d\.]+%)",
        text,
        re.S
    )

    if m:
        result["bm_mfn"] = m.group(1)

    # ======================================
    # PPN
    # ======================================

    m = re.search(
        r"PPN\s*:\s*([\d\.]+%)",
        text
    )

    if m:
        result["ppn"] = m.group(1)

    # ======================================
    # PPH API
    # ======================================

    m = re.search(
        r"PPH\s*:\s*([\d\.]+%)\s*\(API\)",
        text
    )

    if m:
        result["pph_api"] = m.group(1)

    # ======================================
    # PPH NON API
    # ======================================

    m = re.search(
        r"PPH\s*:\s*([\d\.]+%)\s*\(NON-API\)",
        text
    )

    if m:
        result["pph_non_api"] = m.group(1)

    # ======================================
    # IMPORT LICENSES
    # ======================================

    izin_pattern = re.compile(
        r"Nama Izin\s*:\s*(.*?)\s*"
        r"Kode Izin Kepabeanan\s*:\s*(\d+)",
        re.S
    )

    matches = izin_pattern.findall(text)

    for nama, kode in matches:

        result["izin"].append(
            {
                "kode": kode.strip(),
                "nama": nama.strip()
            }
        )

    return result


# ======================================
# LOAD EXCEL
# ======================================

print("Opening workbook...")

wb = load_workbook(FILE_PATH)

ws_tarif = wb["HS Code Tarif"]
ws_izin = wb["HS Code Izin"]

# ======================================
# BUILD PROCESSING QUEUE
# ======================================

hs_queue = []

for row in range(2, ws_tarif.max_row + 1):

    hs_code = ws_tarif[f"A{row}"].value

    bm_mfn = ws_tarif[f"B{row}"].value

    ppn = ws_tarif[f"C{row}"].value

    if hs_code and (
        bm_mfn is None
        or ppn is None
    ):

        hs_queue.append(
            (
                row,
                str(hs_code).strip()
            )
        )

print(
    f"Total HS Codes to process: {len(hs_queue)}"
)

results = []

# ======================================
# PLAYWRIGHT
# ======================================

with sync_playwright() as p:

    browser = p.chromium.launch(
        headless=HEADLESS
    )

    page = browser.new_page()

    for row, hs_code in hs_queue:

        print(
            f"PROCESSING - Row {row} - {hs_code}"
        )

        try:

            data = get_hs_info(
                page,
                hs_code
            )

            results.append(data)

            # ======================================
            # WRITE TARIFF DATA
            # ======================================

            ws_tarif[f"B{row}"] = data["bm_mfn"]

            ws_tarif[f"C{row}"] = data["ppn"]

            ws_tarif[f"D{row}"] = data["pph_api"]

            ws_tarif[f"E{row}"] = data["pph_non_api"]

            # ======================================
            # WRITE LICENSE DATA
            # ======================================

            next_row = ws_izin.max_row + 1

            for izin in data["izin"]:

                ws_izin[f"A{next_row}"] = hs_code

                ws_izin[f"B{next_row}"] = izin["kode"]

                ws_izin[f"C{next_row}"] = izin["nama"]

                next_row += 1

            print(
                f"SUCCESS - {hs_code}"
            )

        except Exception as e:

            print(
                f"FAILED - {hs_code}: {e}"
            )

    browser.close()

# ======================================
# SAVE EXCEL
# ======================================

wb.save(FILE_PATH)

print(
    f"Workbook saved: {FILE_PATH}"
)

# ======================================
# SAVE JSON
# ======================================

with open(
    OUTPUT_JSON,
    "w",
    encoding="utf-8"
) as f:

    json.dump(
        results,
        f,
        indent=4,
        ensure_ascii=False
    )

print(
    f"JSON saved: {OUTPUT_JSON}"
)

# ======================================
# SUMMARY
# ======================================

print(
    f"Processed {len(results)} HS Codes"
)

print(
    "Process completed successfully."
)